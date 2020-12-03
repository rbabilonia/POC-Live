from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
import webbrowser
# Read elements from Xlsx
import openpyxl
path = "/Users/rbabilonia/Documents/Automation/dataurls.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
global listaurl
listaurl = []
for r in range(1,rows+1):
    for c in range(1,cols+1):
        # print(sheet.cell(row=r, column=c).value)
        elemento = (sheet.cell(row=r, column=c).value)
        listaurl.insert(c,elemento)
# print (listaurl[:])
# End of read Xlsx
options = webdriver.ChromeOptions()
options.headless = True
# driver = webdriver.Chrome(executable_path=r"/usr/local/bin/chromedriver")
# driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.implicitly_wait(20)
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--disable-extensions')
chrome_options.add_argument('--profile-directory=Default')
chrome_options.add_argument("--incognito")
chrome_options.add_argument("--disable-plugins-discovery")
chrome_options.add_argument("--start-maximized")
# driver = webdriver.Chrome(chrome_options=chrome_options)
#driver.delete_all_cookies()
# driver.set_window_size(800,800)
# driver.set_window_position(0,0)
ele = listaurl[0]
driver.get(ele)
webbrowser.open(ele)
 # driver.implicitly_wait(20)
# selectbrand = driver.find_element_by_id("makeCode")
# drp = Select(selectbrand)
# drp.deselect_by_visible_text('Jeep')
# selectbrand.send_keys("Jeep")
# selectbrand.clear()
# selectmodel = driver.find_element_by_id("ModelCodeJEEPRENEG")
# selectmodel.send_keys("Renegade")
# selectmodel.clear()
# clickbtn = driver.find_element_by_id("search")
# clickbtn.click()
# driver.get('https://www.amazon.com/')
while True:
    folder = "/Users/rbabilonia/Documents/Automation/FCA"
    time_string = time.asctime().replace(":", " ")
    file_name = folder + time_string + ".png"
    # driver.refresh()
    # driver.implicitly_wait(7)
    # driver.save_screenshot(file_name)
    '''Full Screen Shot'''
    S = lambda X: driver.execute_script('return document.body.parentNode.scroll' + X)
    driver.set_window_size(S('Width'), S('Height'))
    driver.find_element_by_tag_name('body').screenshot(file_name)
    # driver.refresh()
    driver.maximize_window()
    driver.exe
    time.sleep(10)